"""
tests/integration/test_pipeline.py
====================================
Integration tests for the full attendance processing pipeline.

These tests exercise every layer in sequence — just as ``main.py`` does —
but without touching the filesystem for PDF extraction (the OCR step is
stubbed with realistic synthetic text).

Two scenarios are tested end-to-end:
  - TYPE_A pipeline: synthetic TYPE_A OCR text → classify → parse →
    transform → render HTML / Excel (PDF rendering is mocked).
  - TYPE_B pipeline: same flow with TYPE_B text.

Additional edge-case tests:
  - Low-confidence text raises LowConfidenceError before parsing.
  - Unknown report type from a custom factory raises UnknownReportTypeError.
  - Rendered files are valid (HTML parseable, xlsx readable by openpyxl).
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from classification.classifier import Classifier
from errors import LowConfidenceError, UnknownReportTypeError
from generation.excel_renderer import ExcelRenderer
from generation.html_renderer import HtmlRenderer
from generation.pdf_renderer import PdfRenderer
from parsers.parser_factory import ParserFactory
from transformation.service import TransformationService


# ---------------------------------------------------------------------------
# Realistic synthetic OCR text (matches the compiled regex patterns)
# ---------------------------------------------------------------------------

_TYPE_A_OCR = """\
נ.ע. הנשר כח אדם בע"מ
תאריך  יום  מיקום  כניסה  יציאה  הפסקה  סה"כ  100%  125%  150%  שבת
01/01/24 יום ראשון מפעל 08:00 17:00 00:30 8.5 8.0 0.5 0.0 0.0
02/01/24 יום שני מפעל 08:10 16:50 00:30 8.2 8.0 0.2 0.0 0.0
03/01/24 יום שלישי מפעל 07:50 17:10 00:30 8.8 8.0 0.8 0.0 0.0
07/01/24 יום ראשון מפעל 08:00 18:30 00:30 9.5 8.0 1.0 0.5 0.0
08/01/24 יום שני מפעל 08:05 17:05 00:30 8.5 8.0 0.5 0.0 0.0
5 0 43.5 40.0 3.0 0.5 0.0
נסיעות 350
"""

_TYPE_B_OCR = """\
סה"כ ימי עבודה לחודש   5
סה"כ שעות חודשיות      45.0
מחיר לשעה              35.5
סה"כ לתשלום            1597.5
תאריך  יום  כניסה  יציאה  סה"כ  הערות
| 01/01/24 | ראשון | 08:00 | 17:00 | 9.0
| 02/01/24 | שני   | 08:15 | 17:15 | 9.0
| 03/01/24 | שלישי | 07:50 | 16:50 | 9.0
| 07/01/24 | ראשון | 08:00 | 17:00 | 9.0
| 08/01/24 | שני   | 08:05 | 17:05 | 9.0
"""

# Text with one keyword from each side — should produce very low confidence
_AMBIGUOUS_OCR = "125% שעות חודשיות"


# ---------------------------------------------------------------------------
# Pipeline runner helpers
# ---------------------------------------------------------------------------

def _run_pipeline(ocr_text: str, output_dir: Path, render_pdf: bool = False):
    """
    Execute classify → parse → transform → render for *ocr_text*.

    Returns (report_after_transform, rendered_paths).
    """
    classifier = Classifier(confidence_threshold=0.1)
    factory    = ParserFactory()
    service    = TransformationService()

    result = classifier.classify(ocr_text)
    parser = factory.get_parser(result.report_type)
    report = parser.parse(ocr_text, source_file="synthetic.pdf")
    report = service.transform(report)

    rendered: dict[str, Path] = {}
    rendered["html"] = HtmlRenderer().render(report, output_dir)
    rendered["xlsx"] = ExcelRenderer().render(report, output_dir)

    if render_pdf:
        mock_weasy = MagicMock(return_value=MagicMock())
        with patch.dict(sys.modules, {"weasyprint": MagicMock(HTML=mock_weasy)}):
            rendered["pdf"] = PdfRenderer().render(report, output_dir)

    return report, rendered


# ===========================================================================
# TYPE_A end-to-end
# ===========================================================================

class TestTypeAPipeline:
    def test_classifier_identifies_type_a(self):
        clf = Classifier(confidence_threshold=0.1)
        result = clf.classify(_TYPE_A_OCR)
        assert result.report_type == "TYPE_A"

    def test_parse_extracts_five_rows(self):
        clf = Classifier(confidence_threshold=0.1)
        result = clf.classify(_TYPE_A_OCR)
        parser = ParserFactory().get_parser(result.report_type)
        report = parser.parse(_TYPE_A_OCR)
        assert len(report.rows) == 5

    def test_transform_preserves_row_count(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert len(report.rows) == 5

    def test_transform_all_clocks_valid(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_A_OCR, tmp_path)
        for row in report.rows:
            assert row.clock.exit > row.clock.entry

    def test_transform_entry_within_bounds(self, tmp_path):
        from config.rules import RULES_REGISTRY
        report, _ = _run_pipeline(_TYPE_A_OCR, tmp_path)
        wb = RULES_REGISTRY["TYPE_A"].workday
        for row in report.rows:
            assert wb.earliest_entry <= row.clock.entry <= wb.latest_entry

    def test_html_file_created(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert rendered["html"].exists()
        assert rendered["html"].suffix == ".html"

    def test_html_contains_report_type(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        content = rendered["html"].read_text(encoding="utf-8")
        assert "TYPE_A" in content

    def test_html_contains_all_five_rows(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        content = rendered["html"].read_text(encoding="utf-8")
        # Each row has a <tr>; plus header → at least 6 <tr> elements
        assert content.count("<tr>") >= 6

    def test_xlsx_file_created(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert rendered["xlsx"].exists()
        assert rendered["xlsx"].suffix == ".xlsx"

    def test_xlsx_has_correct_sheet_names(self, tmp_path):
        import openpyxl
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        wb = openpyxl.load_workbook(str(rendered["xlsx"]))
        assert "נוכחות" in wb.sheetnames
        assert "סיכום"  in wb.sheetnames

    def test_xlsx_data_rows_count(self, tmp_path):
        import openpyxl
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path)
        wb = openpyxl.load_workbook(str(rendered["xlsx"]))
        ws = wb["נוכחות"]
        # 1 header + 5 data + 1 totals = 7
        assert ws.max_row == 7

    def test_summary_total_days_equals_row_count(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert report.summary.total_days == len(report.rows)

    def test_total_hours_positive(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert report.summary.total_hours > 0


# ===========================================================================
# TYPE_B end-to-end
# ===========================================================================

class TestTypeBPipeline:
    def test_classifier_identifies_type_b(self):
        clf = Classifier(confidence_threshold=0.1)
        result = clf.classify(_TYPE_B_OCR)
        assert result.report_type == "TYPE_B"

    def test_parse_extracts_five_rows(self):
        clf = Classifier(confidence_threshold=0.1)
        result = clf.classify(_TYPE_B_OCR)
        parser = ParserFactory().get_parser(result.report_type)
        report = parser.parse(_TYPE_B_OCR)
        assert len(report.rows) == 5

    def test_transform_all_clocks_valid(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_B_OCR, tmp_path)
        for row in report.rows:
            assert row.clock.exit > row.clock.entry

    def test_html_contains_type_b(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_B_OCR, tmp_path)
        content = rendered["html"].read_text(encoding="utf-8")
        assert "TYPE_B" in content

    def test_xlsx_summary_sheet_contains_hourly_rate(self, tmp_path):
        import openpyxl
        _, rendered = _run_pipeline(_TYPE_B_OCR, tmp_path)
        wb = openpyxl.load_workbook(str(rendered["xlsx"]))
        ws = wb["סיכום"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("תעריף" in str(lbl) for lbl in labels if lbl)

    def test_summary_hourly_rate_preserved(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_B_OCR, tmp_path)
        assert report.summary.hourly_rate == pytest.approx(35.5)

    def test_summary_total_pay_preserved(self, tmp_path):
        report, _ = _run_pipeline(_TYPE_B_OCR, tmp_path)
        assert report.summary.total_pay == pytest.approx(1597.5)


# ===========================================================================
# PDF rendering (mocked WeasyPrint)
# ===========================================================================

class TestPdfRendering:
    def test_pdf_file_path_returned(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_A_OCR, tmp_path, render_pdf=True)
        assert rendered["pdf"].suffix == ".pdf"
        assert rendered["pdf"].parent == tmp_path

    def test_weasyprint_called_once(self, tmp_path):
        mock_weasy_instance = MagicMock()
        mock_weasy_cls = MagicMock(return_value=mock_weasy_instance)

        clf     = Classifier(confidence_threshold=0.1)
        result  = clf.classify(_TYPE_A_OCR)
        parser  = ParserFactory().get_parser(result.report_type)
        report  = parser.parse(_TYPE_A_OCR)
        report  = TransformationService().transform(report)

        with patch.dict(sys.modules, {"weasyprint": MagicMock(HTML=mock_weasy_cls)}):
            PdfRenderer().render(report, tmp_path)

        mock_weasy_instance.write_pdf.assert_called_once()


# ===========================================================================
# Error / edge-case scenarios
# ===========================================================================

class TestPipelineEdgeCases:
    def test_low_confidence_text_raises_before_parsing(self):
        clf = Classifier(confidence_threshold=0.5)
        with pytest.raises(LowConfidenceError):
            clf.classify(_AMBIGUOUS_OCR)

    def test_empty_ocr_raises_low_confidence(self):
        clf = Classifier()
        with pytest.raises(LowConfidenceError):
            clf.classify("")

    def test_custom_factory_unknown_type_raises(self):
        clf = Classifier(confidence_threshold=0.1)
        result = clf.classify(_TYPE_A_OCR)

        # Factory that only knows TYPE_B
        factory = ParserFactory(registry={
            "TYPE_B": ParserFactory().get_parser("TYPE_B"),
        })
        with pytest.raises(UnknownReportTypeError):
            factory.get_parser(result.report_type)  # TYPE_A not registered

    def test_all_three_output_files_created(self, tmp_path):
        _, rendered = _run_pipeline(_TYPE_B_OCR, tmp_path, render_pdf=True)
        for fmt in ("html", "xlsx", "pdf"):
            assert fmt in rendered
            assert rendered[fmt].parent == tmp_path

    def test_multiple_runs_produce_separate_files(self, tmp_path):
        """Two sequential renders should not overwrite each other (timestamps differ)."""
        import time as _time
        _, r1 = _run_pipeline(_TYPE_A_OCR, tmp_path)
        _time.sleep(1)  # ensure different timestamp in filename
        _, r2 = _run_pipeline(_TYPE_A_OCR, tmp_path)
        assert r1["html"] != r2["html"]
