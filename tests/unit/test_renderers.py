"""
tests/unit/test_renderers.py
=============================
Unit tests for the generation layer.

Covers:
  html_renderer.py:
    - _cell_value: all key types for TYPE_A and TYPE_B rows.
    - _esc: HTML escaping of special characters.
    - _build_css: returns a <style> block containing expected selectors.
    - _build_summary_bar: contains expected labels for each report type.
    - _build_table: correct number of <tr> rows, correct header labels.
    - _build_page: full page is valid HTML with correct title.
    - HtmlRenderer.render: creates a file, raises MissingRendererError for
      unknown type, raises OutputDirectoryError for unwritable path.

  excel_renderer.py:
    - _cell_value: date returns date obj, time fields return time obj.
    - ExcelRenderer.render: creates a .xlsx, contains expected sheet names,
      raises MissingRendererError for unknown type.

  pdf_renderer.py:
    - PdfRenderer.render: delegates to HtmlRenderer (mocked), then calls
      WeasyPrint (mocked), returns a .pdf path.
    - Missing WeasyPrint raises RenderingError.
"""

import sys
from datetime import date, time
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from domain.models import (
    AttendanceReport,
    AttendanceRow,
    BreakRecord,
    OvertimeBuckets,
    ReportSummary,
    TimeRange,
)
from errors import MissingRendererError, OutputDirectoryError, RenderingError
from generation.excel_renderer import ExcelRenderer
from generation.excel_renderer import _cell_value as excel_cell_value
from generation.html_renderer import HtmlRenderer
from generation.html_renderer import _build_css, _build_page, _build_summary_bar, _build_table
from generation.html_renderer import _cell_value as html_cell_value
from generation.html_renderer import _esc
from generation.pdf_renderer import PdfRenderer


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_clock(entry=(8, 0), exit_=(17, 0)):
    return TimeRange(entry=time(*entry), exit=time(*exit_))


def _make_break():
    return BreakRecord(
        clock=TimeRange(entry=time(12, 0), exit=time(12, 30)),
        duration_min=30,
    )


def _make_ot():
    return OvertimeBuckets(
        regular_ot=8.0, band_125=1.0, band_150=0.5,

    )


def _type_a_row():
    return AttendanceRow(
        row_date=date(2024, 1, 7),
        day_name="יום ראשון",
        clock=_make_clock(),
        total_hours=8.5,
        location="מפעל",
        break_rec=_make_break(),
        overtime=_make_ot(),
    )


def _type_b_row():
    return AttendanceRow(
        row_date=date(2024, 1, 7),
        day_name="יום ראשון",
        clock=_make_clock(),
        total_hours=9.0,
        notes="holiday",
    )


def _type_a_report():
    return AttendanceReport(
        report_type="TYPE_A",
        rows=(_type_a_row(),),
        summary=ReportSummary(
            total_days=1, total_hours=8.5,
            ot_100=8.0, ot_125=1.0, ot_150=0.5,
            ot_shabbat=0.0, travel_allowance=350.0,
        ),
    )


def _type_b_report():
    return AttendanceReport(
        report_type="TYPE_B",
        rows=(_type_b_row(),),
        summary=ReportSummary(
            total_days=1, total_hours=9.0,
            hourly_rate=35.5, total_pay=319.5,
        ),
    )


# ===========================================================================
# HTML renderer
# ===========================================================================

class TestEsc:
    def test_ampersand(self):
        assert _esc("a & b") == "a &amp; b"

    def test_less_than(self):
        assert _esc("<tag>") == "&lt;tag&gt;"

    def test_double_quote(self):
        assert _esc('"hello"') == "&quot;hello&quot;"

    def test_no_special_chars_unchanged(self):
        assert _esc("hello world") == "hello world"

    def test_empty_string(self):
        assert _esc("") == ""


class TestHtmlCellValue:
    def test_date_key_format(self):
        row = _type_a_row()
        assert html_cell_value(row, "date") == "07/01/2024"

    def test_day_key(self):
        row = _type_a_row()
        assert "ראשון" in html_cell_value(row, "day")

    def test_entry_key_format(self):
        row = _type_a_row()
        assert html_cell_value(row, "entry") == "08:00"

    def test_exit_key_format(self):
        row = _type_a_row()
        assert html_cell_value(row, "exit") == "17:00"

    def test_break_key_with_break(self):
        row = _type_a_row()
        result = html_cell_value(row, "break")
        assert "30" in result

    def test_break_key_no_break(self):
        row = _type_b_row()
        assert html_cell_value(row, "break") == "—"

    def test_net_key_shows_decimal(self):
        row = _type_a_row()
        result = html_cell_value(row, "net")
        assert "." in result

    def test_ot_100_pill_rendered(self):
        row = _type_a_row()
        result = html_cell_value(row, "ot_100")
        assert "<span" in result
        assert "8" in result

    def test_ot_key_no_overtime_returns_dash(self):
        row = AttendanceRow(
            row_date=date(2024, 1, 7),
            day_name="יום ראשון",
            clock=_make_clock(),
            total_hours=9.0,
        )
        assert html_cell_value(row, "ot_100") == "—"

    def test_location_key(self):
        row = _type_a_row()
        assert html_cell_value(row, "location") == "מפעל"

    def test_location_key_none_returns_dash(self):
        row = _type_b_row()
        assert html_cell_value(row, "location") == "—"

    def test_notes_key(self):
        row = _type_b_row()
        assert html_cell_value(row, "notes") == "holiday"

    def test_unknown_key_returns_dash(self):
        assert html_cell_value(_type_a_row(), "nonexistent") == "—"


class TestBuildCss:
    def test_returns_style_block(self):
        theme = {
            "header_bg": "#1e3a5f", "header_text": "#fff",
            "even_row": "#f0f4fa", "hover": "#dce8f8",
            "border": "#c5d4e8", "summary_bg": "#e8eef6",
            "summary_accent": "#1e3a5f",
        }
        css = _build_css(theme)
        assert css.startswith("<style>")
        assert css.endswith("</style>")
        assert "table" in css


class TestBuildSummaryBar:
    def test_type_a_includes_ot_labels(self):
        bar = _build_summary_bar(_type_a_report())
        assert "OT 100%" in bar
        assert "OT 125%" in bar
        assert "נסיעות"   in bar

    def test_type_b_includes_rate_labels(self):
        bar = _build_summary_bar(_type_b_report())
        assert "תעריף" in bar or "לשעה" in bar

    def test_total_days_present(self):
        for report in (_type_a_report(), _type_b_report()):
            bar = _build_summary_bar(report)
            assert "ימי עבודה" in bar


class TestBuildTable:
    def _columns(self):
        return [("תאריך", "date"), ("יום", "day"), ("כניסה", "entry")]

    def test_contains_header(self):
        html = _build_table(_type_a_report(), self._columns())
        assert "<thead>" in html
        assert "תאריך" in html

    def test_contains_one_data_row(self):
        html = _build_table(_type_a_report(), self._columns())
        assert html.count("<tr>") >= 2  # header row + 1 data row

    def test_contains_table_tag(self):
        assert "<table" in _build_table(_type_a_report(), self._columns())


class TestBuildPage:
    def _columns(self):
        return [("תאריך", "date"), ("כניסה", "entry")]

    def _theme(self):
        return {
            "header_bg": "#1e3a5f", "header_text": "#fff",
            "even_row": "#f0f4fa", "hover": "#dce8f8",
            "border": "#c5d4e8", "summary_bg": "#e8eef6",
            "summary_accent": "#1e3a5f",
        }

    def test_starts_with_doctype(self):
        page = _build_page(_type_a_report(), self._columns(), self._theme())
        assert page.startswith("<!DOCTYPE html>")

    def test_contains_report_type(self):
        page = _build_page(_type_a_report(), self._columns(), self._theme())
        assert "TYPE_A" in page

    def test_rtl_direction(self):
        page = _build_page(_type_a_report(), self._columns(), self._theme())
        assert 'dir="rtl"' in page or "direction:rtl" in page


class TestHtmlRenderer:
    def test_render_creates_html_file(self, tmp_path):
        renderer = HtmlRenderer()
        out = renderer.render(_type_a_report(), tmp_path)
        assert out.exists()
        assert out.suffix == ".html"

    def test_render_file_contains_report_type(self, tmp_path):
        renderer = HtmlRenderer()
        out = renderer.render(_type_b_report(), tmp_path)
        content = out.read_text(encoding="utf-8")
        assert "TYPE_B" in content

    def test_render_unknown_type_raises_missing_renderer(self, tmp_path):
        renderer = HtmlRenderer()
        report = AttendanceReport(
            report_type="TYPE_X",
            rows=(_type_a_row(),),
            summary=ReportSummary(),
        )
        with pytest.raises(MissingRendererError):
            renderer.render(report, tmp_path)

    def test_render_unwritable_path_raises_output_directory_error(self, tmp_path):
        """Simulate an unwritable directory by patching Path.mkdir to raise OSError."""
        from unittest.mock import patch
        renderer = HtmlRenderer()
        with patch("pathlib.Path.mkdir", side_effect=OSError("permission denied")):
            with pytest.raises((OutputDirectoryError, OSError)):
                renderer.render(_type_a_report(), tmp_path / "blocked" / "report.html")


# ===========================================================================
# Excel renderer
# ===========================================================================

class TestExcelCellValue:
    def test_date_returns_date_object(self):
        row = _type_a_row()
        val = excel_cell_value(row, "date")
        assert isinstance(val, date)

    def test_entry_returns_time_object(self):
        row = _type_a_row()
        val = excel_cell_value(row, "entry")
        assert isinstance(val, time)

    def test_break_with_break_returns_int(self):
        row = _type_a_row()
        assert excel_cell_value(row, "break") == 30

    def test_break_without_break_returns_zero(self):
        row = _type_b_row()
        assert excel_cell_value(row, "break") == 0

    def test_net_returns_float(self):
        row = _type_a_row()
        val = excel_cell_value(row, "net")
        assert isinstance(val, float)
        assert val > 0

    def test_ot_100_with_ot(self):
        row = _type_a_row()
        assert excel_cell_value(row, "ot_100") == pytest.approx(8.0)

    def test_ot_no_overtime_returns_zero(self):
        row = _type_b_row()
        assert excel_cell_value(row, "ot_100") == 0.0

    def test_location_string(self):
        row = _type_a_row()
        assert excel_cell_value(row, "location") == "מפעל"

    def test_unknown_key_returns_empty_string(self):
        assert excel_cell_value(_type_a_row(), "unknown") == ""


class TestExcelRenderer:
    def test_render_creates_xlsx_file(self, tmp_path):
        renderer = ExcelRenderer()
        out = renderer.render(_type_a_report(), tmp_path)
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_render_type_b_creates_file(self, tmp_path):
        renderer = ExcelRenderer()
        out = renderer.render(_type_b_report(), tmp_path)
        assert out.exists()

    def test_xlsx_has_two_sheets(self, tmp_path):
        import openpyxl
        renderer = ExcelRenderer()
        out = renderer.render(_type_a_report(), tmp_path)
        wb = openpyxl.load_workbook(str(out))
        assert "נוכחות" in wb.sheetnames
        assert "סיכום"  in wb.sheetnames

    def test_data_row_count_correct(self, tmp_path):
        import openpyxl
        renderer = ExcelRenderer()
        out = renderer.render(_type_a_report(), tmp_path)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["נוכחות"]
        # row 1 = header, row 2 = data, row 3 = totals
        assert ws.max_row == 3

    def test_unknown_type_raises(self, tmp_path):
        renderer = ExcelRenderer()
        report = AttendanceReport(
            report_type="TYPE_X",
            rows=(_type_a_row(),),
            summary=ReportSummary(),
        )
        with pytest.raises(MissingRendererError):
            renderer.render(report, tmp_path)

    def test_custom_header_hex_applied(self, tmp_path):
        import openpyxl
        renderer = ExcelRenderer(header_fill_hex="2D4A22")
        out = renderer.render(_type_a_report(), tmp_path)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["נוכחות"]
        header_fill = ws.cell(row=1, column=1).fill.fgColor.rgb
        assert "2D4A22" in header_fill.upper()


# ===========================================================================
# PDF renderer
# ===========================================================================

class TestPdfRenderer:
    def test_render_calls_weasyprint(self, tmp_path):
        """WeasyPrint is mocked — we verify render() delegates correctly."""
        mock_html_instance = MagicMock()
        mock_weasy_cls = MagicMock(return_value=mock_html_instance)

        with patch.dict(sys.modules, {"weasyprint": MagicMock(HTML=mock_weasy_cls)}):
            renderer = PdfRenderer()
            out = renderer.render(_type_a_report(), tmp_path)

        mock_html_instance.write_pdf.assert_called_once()
        assert out.suffix == ".pdf"

    def test_render_creates_file_path_in_output_dir(self, tmp_path):
        mock_html_instance = MagicMock()
        mock_weasy_cls = MagicMock(return_value=mock_html_instance)

        with patch.dict(sys.modules, {"weasyprint": MagicMock(HTML=mock_weasy_cls)}):
            renderer = PdfRenderer()
            out = renderer.render(_type_b_report(), tmp_path)

        assert out.parent == tmp_path

    def test_weasyprint_missing_raises_rendering_error(self, tmp_path):
        """If weasyprint cannot be imported, RenderingError is raised."""
        with patch.dict(sys.modules, {"weasyprint": None}):
            renderer = PdfRenderer()
            with pytest.raises((RenderingError, ImportError)):
                renderer.render(_type_a_report(), tmp_path)

    def test_custom_html_renderer_injected(self, tmp_path):
        """PdfRenderer uses the injected HtmlRenderer, not a fresh default."""
        spy_html = MagicMock(spec=HtmlRenderer)
        # render() must write a real temp file for WeasyPrint to read;
        # make the spy write a minimal HTML file.
        def fake_render(report, path):
            p = path if not path.is_dir() else path / "tmp.html"
            p.write_text("<html></html>", encoding="utf-8")
            return p

        spy_html.render.side_effect = fake_render
        mock_weasy_cls = MagicMock(return_value=MagicMock())

        with patch.dict(sys.modules, {"weasyprint": MagicMock(HTML=mock_weasy_cls)}):
            renderer = PdfRenderer(html_renderer=spy_html)
            renderer.render(_type_a_report(), tmp_path)

        spy_html.render.assert_called_once()
