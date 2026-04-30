"""
generation/excel_renderer.py
=============================
Renders an AttendanceReport to an Excel (.xlsx) file using openpyxl.

Two sheets:
  - "נוכחות"  — per-row attendance data
  - "סיכום"   — summary key/value pairs
"""

from __future__ import annotations

import logging
from datetime import date, datetime, time
from pathlib import Path

from domain.errors import MissingRendererError, OutputDirectoryError
from domain.models import AttendanceReport, AttendanceRow
from generation.base import BaseRenderer

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column specs  — (header_label, extractor_key)
# ---------------------------------------------------------------------------

_COLUMNS_TYPE_A: list[tuple[str, str]] = [
    ("תאריך",     "date"),
    ("יום",       "day"),
    ("כניסה",     "entry"),
    ("יציאה",     "exit"),
    ("הפסקה",     "break"),
    ("שעות נטו", "net"),
    ("OT 100%",   "ot_100"),
    ("OT 125%",   "ot_125"),
    ("OT 150%",   "ot_150"),
    ("מקום",      "location"),
    ("שבת",       "shabbat"),
]

_COLUMNS_TYPE_B: list[tuple[str, str]] = [
    ("תאריך",     "date"),
    ("יום",       "day"),
    ("כניסה",     "entry"),
    ("יציאה",     "exit"),
    ("שעות נטו", "net"),
    ("הערות",     "notes"),
]

_DEFAULT_COLUMN_MAP: dict[str, list[tuple[str, str]]] = {
    "TYPE_A": _COLUMNS_TYPE_A,
    "TYPE_B": _COLUMNS_TYPE_B,
}


# ---------------------------------------------------------------------------
# Cell-value extractor  — returns native Python types for Excel
# ---------------------------------------------------------------------------

def _cell_value(row: AttendanceRow, key: str):
    if key == "date":
        return row.row_date
    if key == "day":
        return row.day_name
    if key == "entry":
        return row.clock.entry
    if key == "exit":
        return row.clock.exit
    if key == "break":
        return row.break_rec.duration_min if row.break_rec else 0
    if key == "net":
        return round(row.net_hours, 2)
    if key in ("ot_100", "ot_125", "ot_150", "shabbat"):
        if row.overtime is None:
            return 0.0
        attr = {"ot_100": "regular_ot", "ot_125": "band_125",
                "ot_150": "band_150", "shabbat": "weekend_ot"}[key]
        return getattr(row.overtime, attr)
    if key == "location":
        return row.location or ""
    if key == "notes":
        return row.notes or ""
    return ""


# ---------------------------------------------------------------------------
# Summary rows  — list of (label, value) pairs per report type
# ---------------------------------------------------------------------------

def _summary_rows(report: AttendanceReport) -> list[tuple[str, object]]:
    s = report.summary
    rows: list[tuple[str, object]] = []
    if s.total_days  is not None: rows.append(("ימי עבודה",    s.total_days))
    if s.total_hours is not None: rows.append(('סה"כ שעות',   s.total_hours))
    if report.report_type == "TYPE_A":
        for lbl, val in [
            ("OT 100%",  s.ot_100),
            ("OT 125%",  s.ot_125),
            ("OT 150%",  s.ot_150),
            ("שבת",      s.ot_shabbat),
            ("נסיעות",   s.travel_allowance),
            ("בונוס",    s.bonus),
        ]:
            if val is not None:
                rows.append((lbl, val))
    elif report.report_type == "TYPE_B":
        for lbl, val in [
            ("תעריף לשעה",   s.hourly_rate),
            ('סה"כ לתשלום', s.total_pay),
        ]:
            if val is not None:
                rows.append((lbl, val))
    return rows


# ---------------------------------------------------------------------------
# Renderer
# ---------------------------------------------------------------------------

class ExcelRenderer(BaseRenderer):
    """Renders an AttendanceReport to a .xlsx file."""

    def __init__(
        self,
        column_map:      dict[str, list[tuple[str, str]]] | None = None,
        header_fill_hex: str = "1E3A5F",
    ) -> None:
        self._column_map      = column_map if column_map is not None else _DEFAULT_COLUMN_MAP
        self._header_fill_hex = header_fill_hex

    def render(self, report: AttendanceReport, output_path: Path) -> Path:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ImportError("openpyxl is required for ExcelRenderer") from exc

        columns = self._column_map.get(report.report_type)
        if columns is None:
            raise MissingRendererError(
                report_type=report.report_type,
                registered=list(self._column_map),
            )

        self._ensure_output_dir(output_path)

        default_name = (
            f"attendance_{report.report_type.lower()}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        dest = self._resolve_output_path(output_path, default_name)
        # Ensure .xlsx extension
        if dest.suffix != ".xlsx":
            dest = dest.with_suffix(".xlsx")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Attendance ──────────────────────────────────────────
        ws_data = wb.active
        ws_data.title = "נוכחות"

        header_fill = PatternFill("solid", fgColor=self._header_fill_hex.upper())
        header_font = Font(bold=True, color="FFFFFF")

        # Header row
        for col_idx, (label, _) in enumerate(columns, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="right")

        # Data rows
        for row_idx, row in enumerate(report.rows, start=2):
            for col_idx, (_, key) in enumerate(columns, start=1):
                ws_data.cell(row=row_idx, column=col_idx, value=_cell_value(row, key))

        # Totals row
        totals_row_idx = len(report.rows) + 2
        numeric_keys = {"net", "ot_100", "ot_125", "ot_150", "shabbat"}
        for col_idx, (_, key) in enumerate(columns, start=1):
            if col_idx == 1:
                ws_data.cell(row=totals_row_idx, column=1, value='סה"כ')
            elif key in numeric_keys:
                total = sum(_cell_value(r, key) for r in report.rows)
                ws_data.cell(row=totals_row_idx, column=col_idx, value=round(total, 2))

        # ── Sheet 2: Summary ─────────────────────────────────────────────
        ws_summary = wb.create_sheet(title="סיכום")
        for r_idx, (label, value) in enumerate(_summary_rows(report), start=1):
            ws_summary.cell(row=r_idx, column=1, value=label)
            ws_summary.cell(row=r_idx, column=2, value=value)

        try:
            wb.save(str(dest))
        except OSError as exc:
            raise OutputDirectoryError(dest, str(exc)) from exc

        logger.debug("ExcelRenderer.render: written → %s", dest)
        return dest
