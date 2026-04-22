import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.processing.rendering.base_renderer import BaseRenderer
from app.templates import TEMPLATES
from app.row_extractor import get_cell_value
from app.processing.rules.base_rules import minutes_to_str

HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
SUMMARY_FILL = PatternFill("solid", fgColor="D5E8D4")
INFO_FILL    = PatternFill("solid", fgColor="EBF5FB")
ALT_FILL     = PatternFill("solid", fgColor="F2F3F4")
SHABAT_FILL  = PatternFill("solid", fgColor="A0A0A0")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(bold=True, size=10)
LABEL_FONT   = Font(bold=True, size=10, color="2C3E50")
NORMAL_FONT  = Font(size=10)
SHABAT_FONT  = Font(size=10, italic=True, color="333333")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
THIN   = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class ExcelRenderer(BaseRenderer):

    def render(self, report, output_path: str) -> str:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        doc_type = report.doc_type
        template = TEMPLATES[doc_type]

        wb = Workbook()
        ws = wb.active
        ws.title = f"נוכחות {doc_type}"
        ws.sheet_view.rightToLeft = True

        info_rows = self._info_rows(report, doc_type)

        if doc_type == "B":
            start_row = self._write_info(ws, info_rows, start=1) + 2
        else:
            start_row = 1

        for col, cdef in enumerate(template, start=1):
            cell = ws.cell(row=start_row, column=col, value=cdef.header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = cdef.width_xl

        for row_i, line in enumerate(report.lines, start=start_row + 1):
            is_shabat = getattr(line, "is_shabat", False)
            fill = SHABAT_FILL if is_shabat else (ALT_FILL if row_i % 2 == 0 else PatternFill())
            font = SHABAT_FONT if is_shabat else NORMAL_FONT
            for col, cdef in enumerate(template, start=1):
                val = get_cell_value(line, cdef.key) if not is_shabat or cdef.key in ("date", "day") else ""
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.font = font
                cell.fill = fill
                cell.alignment = CENTER
                cell.border = BORDER

        sr = start_row + len(report.lines) + 1
        for col, cdef in enumerate(template, start=1):
            val = self._summary_val(report, doc_type, cdef.key)
            cell = ws.cell(row=sr, column=col, value=val)
            cell.font = BOLD_FONT
            cell.fill = SUMMARY_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        if doc_type == "A":
            self._write_info(ws, info_rows, start=sr + 2)

        wb.save(output_path)
        return output_path

    def _write_info(self, ws, rows: list, start: int) -> int:
        for i, (label, val) in enumerate(rows):
            r = start + i
            lc = ws.cell(row=r, column=1, value=f"{label}:")
            lc.font = LABEL_FONT
            lc.fill = INFO_FILL
            lc.alignment = RIGHT
            vc = ws.cell(row=r, column=2, value=str(val))
            vc.font = NORMAL_FONT
            vc.fill = INFO_FILL
            vc.alignment = CENTER
        return start + len(rows) - 1

    def _info_rows(self, report, doc_type: str) -> list:
        if doc_type == "A":
            return [
                ("ימים", report.days),
                ('סה"כ שעות', minutes_to_str(report.total_hours)),
                ("100% שעות", minutes_to_str(report.hours_100)),
                ("125% שעות", minutes_to_str(report.hours_125)),
                ("150% שעות", minutes_to_str(report.hours_150)),
                ("שבת 150%", minutes_to_str(report.shabat_total)),
                ("בונוס", f"{report.bonus:.2f}"),
                ("נסיעות", f"{report.nesiot:.2f}"),
            ]
        return [
            ("שם העובד", report.worker_name or ""),
            ('סה"כ ימי עבודה לחודש', report.days),
            ('סה"כ שעות חודשיות', minutes_to_str(report.total_hours)),
            ("מחיר לשעה", f"{report.price_per_hour:.2f}"),
            ('סה"כ לתשלום', f"{report.total_payment:.2f}"),
            ("כרטיס עובד לחודש", report.month or ""),
        ]

    def _summary_val(self, report, doc_type: str, key: str) -> str:
        if key == "date":   return 'סה"כ'
        if key == "total":  return minutes_to_str(report.total_hours)
        if key == "hours_100" and doc_type == "A": return minutes_to_str(report.hours_100)
        if key == "hours_125" and doc_type == "A": return minutes_to_str(report.hours_125)
        if key == "hours_150" and doc_type == "A": return minutes_to_str(report.hours_150)
        return ""
