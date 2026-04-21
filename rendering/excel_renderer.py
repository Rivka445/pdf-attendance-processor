# ===== Excel Renderer =====
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rendering.base_renderer import BaseRenderer
from models.type_a import TypeA
from models.type_b import TypeB
from templates import TEMPLATES
from row_extractor import get_cell_value
from rules.base_rules import minutes_to_str

HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
SUMMARY_FILL = PatternFill("solid", fgColor="D5E8D4")
ALT_FILL     = PatternFill("solid", fgColor="F2F3F4")
SHABAT_FILL  = PatternFill("solid", fgColor="A0A0A0")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)
SHABAT_FONT  = Font(size=10, italic=True, color="333333")
CENTER = Alignment(horizontal="center", vertical="center")
THIN   = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class ExcelRenderer(BaseRenderer):

    def render(self, report, output_path: str) -> str:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        doc_type = "A" if isinstance(report, TypeA) else "B"
        template = TEMPLATES[doc_type]

        wb = Workbook()
        ws = wb.active
        ws.title = f"נוכחות {doc_type}"
        ws.sheet_view.rightToLeft = True

        # כותרות
        for col, cdef in enumerate(template, start=1):
            cell = ws.cell(row=1, column=col, value=cdef.header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = cdef.width_xl

        # שורות נתונים
        for row_i, line in enumerate(report.lines, start=2):
            is_shabat = getattr(line, "is_shabat", False)
            fill = SHABAT_FILL if is_shabat else (ALT_FILL if row_i % 2 == 0 else PatternFill())
            font = SHABAT_FONT if is_shabat else NORMAL_FONT
            for col, cdef in enumerate(template, start=1):
                val = get_cell_value(line, cdef.key) if not is_shabat or cdef.key in ("date", "day") else ""
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.font = font
                cell.fill = fill
                cell.alignment = CENTER
                cell.border = BORDER

        # שורת סיכום
        sr = len(report.lines) + 2
        for col, cdef in enumerate(template, start=1):
            val = self._summary_val(report, doc_type, cdef.key)
            cell = ws.cell(row=sr, column=col, value=val)
            cell.font = BOLD_FONT
            cell.fill = SUMMARY_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        wb.save(output_path)
        return output_path

    def _summary_val(self, report, doc_type: str, key: str) -> str:
        if key == "date":
            return 'סה"כ'
        if key == "total":
            return minutes_to_str(report.total_hours)
        if key == "hours_100" and doc_type == "A":
            return minutes_to_str(report.hours_100)
        if key == "hours_125" and doc_type == "A":
            return minutes_to_str(report.hours_125)
        if key == "hours_150" and doc_type == "A":
            return minutes_to_str(report.hours_150)
        return ""
